import os
import openpyxl as xl
import pandas as pd
from sklearn.cluster import KMeans
import matplotlib.pyplot as plt
from sklearn.metrics import silhouette_score
from sklearn.metrics import calinski_harabasz_score
from sklearn.metrics import davies_bouldin_score


def dataCluster(colomn):
    chart = xl.load_workbook(os.path.dirname(__file__) + '/data/raw_data.xlsx', data_only=True)
    sheet = chart['Sheet1']

    data = pd.read_excel(os.path.dirname(__file__) + '/data/raw_data.xlsx', sheet_name='Sheet1', usecols=colomn)
    data.head()
    """
    # clustering
    y_pre = KMeans(n_clusters=3, random_state=0).fit_predict(data)
    
    # write to Excel
    i = 0
    while i < 8638:
        sheet.cell(row=i+2, column=13).value = y_pre[i]
        i += 1
    chart.save(os.path.dirname(__file__) + '/data/raw_data.xlsx')
    """
    SSE = []
    sil_score = []
    
    # Silhouette Coefficient and SSE for different K
    for k in range(2, 9):
        model = KMeans(n_clusters=k, random_state=0)
        model.fit_predict(data)
        sil_score.append(silhouette_score(data, model.labels_))  # Silhouette Coefficient
        SSE.append(model.inertia_)  # SSE
    
    plt.figure()
    plt.subplot(2, 1, 1)
    plt.plot(range(2, 9), SSE, color='#548235', linestyle='-', marker='.', label="SSE")
    plt.xlabel('k')
    plt.ylabel('SSE')
    plt.subplot(2, 1, 2)
    plt.plot(range(2, 9), sil_score, color='#305496', linestyle='-', marker='.', label="sil_score")
    plt.xlabel('k')
    plt.ylabel('sil_score')
    plt.subplots_adjust(hspace=0.5)
    plt.show()



def boolDataCluster():
    chart = xl.load_workbook(os.path.dirname(__file__) + '/data/raw_data.xlsx', data_only=True)
    sheet = chart['Sheet1']

    # Transfer 4-bit Boolean value to decimal
    for i in range(8638):
        temp = sheet.cell(row=i+2, column=3).value * 2 ** 3 + sheet.cell(row=i+2, column=4).value * 2 ** 2 + sheet.cell(row=i+2, column=5).value * 2 + sheet.cell(row=i+2, column=6).value
        sheet.cell(row=i+2, column=7).value = temp

    # Transform 16 categories into 4 clusters
    for i in range(8638):
        temp = sheet.cell(i+2, 7).value
        if 0 <= temp <= 3:
            sheet.cell(i+2, 7).value = 0
        elif 4 <= temp <= 7:
            sheet.cell(i+2, 7).value = 1
        elif 8 <= temp <= 11:
            sheet.cell(i+2, 7).value = 2
        elif 12 <= temp <= 15:
            sheet.cell(i+2, 7).value = 3

    chart.save(os.path.dirname(__file__) + '/data/raw_data.xlsx')


if __name__ == '__main__':
    # Clustering for continuous attributes
    dataCluster('H:M, Q')

    # Clustering for discrete attributes
    # boolDataCluster()
