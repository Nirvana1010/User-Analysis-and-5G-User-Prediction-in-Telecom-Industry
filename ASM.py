import pandas as pd
import openpyxl as xl
import os
from mlxtend.preprocessing import TransactionEncoder
from mlxtend.frequent_patterns import apriori
from mlxtend.frequent_patterns import association_rules


def dataTransform():
    chart = xl.load_workbook(os.path.dirname(__file__) + '/data/raw_data.xlsx', data_only=True)
    sheet1 = chart['Sheet1']
    sheet5 = chart['Sheet5']

    n = 1
    i = 2
    packages = []
    while i < 8640:
        # For one user with multiple plans
        if sheet1.cell(i, 19).value != 1:
            dup = sheet1.cell(i, 19).value  # the number of plans
            temp = [sheet1.cell(i, 1).value, dup]  # user / plan
            j = 0
            while j < dup:
                temp.append(sheet1.cell(i+j, 2).value)  # type of the plan
                j += 1
            i += dup  # next user
            packages.append(temp)
        else:
            i += 1

    df = pd.DataFrame(packages)
    df.to_excel(os.path.dirname(__file__) + '/data/package_data.xlsx')


def setSort():
    chart = xl.load_workbook(os.path.dirname(__file__) + '/data/raw_data.xlsx', data_only=True)
    sheet5 = chart['Sheet5']

    for i in range(2, 541):
        temp = []
        for j in range(sheet5.cell(i, 2).value):
            temp.append(sheet5.cell(i, j+3).value)

        # sort set
        temp.sort()
        for j in range(sheet5.cell(i, 2).value):
            sheet5.cell(i, j+3).value = temp[j]

    chart.save(os.path.dirname(__file__) + '/data/raw_data.xlsx')


def dataDiscrete(column):
    data = pd.read_excel(os.path.dirname(__file__) + '/data/association_data.xlsx', sheet_name='Sheet1', usecols=column)
    data.head()

    """
    # Data discretization: binning
    data['RevenueTeam'] = pd.cut(data.TotalRevenue, 20, labels=range(20))  # range = 3w
    data.groupby('RevenueTeam').count()
    data['TotalATeam'] = pd.cut(data.TotalArrearage, 20, labels=range(20))  # range = 24w
    data.groupby('TotalATeam').count()
    data['MaxATeam'] = pd.cut(data.MaxArrearageTime, 5, labels=range(5))  # range = 80
    data.groupby('MaxATeam').count()
    data['RegisterTeam'] = pd.cut(data.Registration, 10, labels=range(10))  # range = 7000
    data.groupby('RegisterTeam').count()
    """

    # Quantile-based discretization
    data['RevenueTeam'] = pd.qcut(data.TotalRevenue, 10, labels=range(8), duplicates='drop')
    data['TotalATeam'] = pd.qcut(data.TotalArrearage, 10, labels=range(4), duplicates='drop')
    data['MaxATeam'] = pd.qcut(data.MaxArrearageTime, 10, labels=range(4), duplicates='drop')
    data['RegisterTeam'] = pd.qcut(data.Registration, 10, labels=range(10), duplicates='drop')
    print(data.groupby('RevenueTeam').count())
    print(data.groupby('TotalATeam').count())
    print(data.groupby('MaxATeam').count())
    print(data.groupby('RegisterTeam').count())

    data.to_excel(os.path.dirname(__file__) + '/data/NEWdiscrete_data.xlsx')


def dataRename():
    result = pd.read_excel(os.path.dirname(__file__) + '/data/NEWdiscrete_data.xlsx', sheet_name='Sheet1')
    result.head()

    # result['RevenueTeam'] = result.RevenueTeam + 'Rev'
    result["RevenueTeamNew"] = ['Rev%i' % i for i in result["RevenueTeam"]]
    result["TotalATeamNew"] = ['TA%i' % i for i in result["TotalATeam"]]
    result["MaxATeamNew"] = ['MA%i' % i for i in result["MaxATeam"]]
    result["RegisterTeamNew"] = ['Reg%i' % i for i in result["RegisterTeam"]]
    result.to_excel(os.path.dirname(__file__) + '/data/discrete_data.xlsx')
    # print(data['RevenueTeam'])


def asssociationRule(open_name, save_name, sheet_name, data_range, s, c, column=None):
    # load data
    data = pd.read_excel(os.path.dirname(__file__) + '/data/%s.xlsx' % open_name, sheet_name=sheet_name)
    data.head()

    # generate frequent itemset
    i = 0
    itemset = []

    if data_range == 'High':
        data = data.where(data.Level >= 4)
        data = data.dropna()
        data = data.reset_index(drop=True)
    elif data_range == 'Low':
        data = data.where(data.Level < 4)
        data = data.dropna()
        data = data.reset_index(drop=True)

    while i < data.shape[0]:
        temp = []
        # association rules between different plans
        if open_name == 'raw_data':
            for j in range(data['PackageNumber'][i]):
                temp.extend([data['Package' + str(j+1)][i]])  # npArray -> list
            itemset.append(temp)
        elif open_name == 'association_data':
            for j in range(len(column)):
                temp.extend([data[column[j]][i]])
            itemset.append(temp)
        elif open_name == 'level_data':
            for j in range(len(column)):
                temp.extend([data[column[j]][i]])
            itemset.append(temp)
        i += 1
    # print(itemset)

    # transfer the itemset to fit model
    TE = TransactionEncoder()
    set_transform = TE.fit_transform(itemset)
    set_df = pd.DataFrame(set_transform, columns=TE.columns_)

    # get frequent itemset
    frequent_itemset = apriori(set_df, min_support=s, use_colnames=True)
    # get association rules
    rules = association_rules(frequent_itemset, metric='confidence',
                              min_threshold=c)
    # set minimum lift
    rules = rules.drop(rules[rules.lift < 1.0].index)
    # print
    rules = rules[['antecedents', 'consequents', 'support',
                   'confidence', 'lift']]
    rules.to_excel(os.path.dirname(__file__) + '/data/%s.xlsx' % save_name)
    print(rules)


if __name__ == '__main__':
    # transform user data and plans into itemsets
    # dataTransform()

    # sort itemsets
    # setSort()

    # data discretization
    # dataDiscrete('H, K, M, O')

    # Rename variable
    # dataRename()

    # index for ASM
    columnIndex = ['Package', 'RegisterTeam']

    # whole users
    asssociationRule('association_data', 'LASM_r', 'Sheet1', 'Low', 0.01, 0.5, columnIndex)

    # high-level users
    # asssociationRule('level_data', 'LASM_pAndRevenue', 'LowLevel', 0.01, 0.2, columnIndex)
