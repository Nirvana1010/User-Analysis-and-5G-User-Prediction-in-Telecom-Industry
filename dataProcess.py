import xlrd
import openpyxl as xl
import os
from scipy import stats
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from sklearn import preprocessing
import time


# 合并两表的属性，对于主键相同的项，对欠费金额、收入加和
def attrMerge():
    """
    Merge 2 tables
    For data points with same primary key, add outstanding fees & revenues
    """
    main_chart = xl.load_workbook(os.path.dirname(__file__) + '/data/raw_data.xlsx', data_only=True)
    main_sheet = main_chart['Sheet1']
    supple_chart = xlrd.open_workbook(os.path.dirname(__file__) + '/data/attr_supplement.xlsx')
    supple_sheet = supple_chart.sheet_by_name('Sheet1')
    print("open 2 files successfully")

    # set primary key
    key_supple = supple_sheet.col_values(5)

    i = 2  # 原表
    while i < 8640:
        if main_sheet.cell(row=i, column=28).value != 1:
            # get repetitive keys and number of repetition
            key_dup = main_sheet.cell(row=i, column=26).value
            dup = main_sheet.cell(row=i, column=28).value
            # print(key_dup, ", ", dup)

            # Calculate outstanding fees & revenues
            """
            total = 0
            j = 1  # 属性表
            while 1:
                if key_dup == key_supple[j]:
                    # print(supple_sheet.cell(j, 5))
                    k = 0
                    while k < dup:
                        total += supple_sheet.cell(j+k, 9).value
                        k += 1
                    # print("total: ", total)
                    break
                j += 1
            # print("第%s行" % (j + 1), "相同")
            main_sheet.cell(row=i, column=30).value = total
            """

            # Calculate maximum length of outstanding charges
            max = 0
            j = 1  # 属性表
            while 1:
                if key_dup == key_supple[j]:
                    # print(supple_sheet.cell(j, 5))
                    k = 0
                    while k < dup:
                        if max < supple_sheet.cell(j + k, 10).value:
                            max = supple_sheet.cell(j + k, 10).value
                        k += 1
                    break
                j += 1
            main_sheet.cell(row=i, column=31).value = max
        else:
            main_sheet.cell(row=i, column=31).value = 0

        if i % 500 == 0:
            print("计算到第%d行" % i)
            main_chart.save(os.path.dirname(__file__) + '/data/raw_data.xlsx')
        i += 1

    main_chart.save(os.path.dirname(__file__) + '/data/raw_data.xlsx')


def boolMake():
    chart = xl.load_workbook(os.path.dirname(__file__) + '/data/raw_data.xlsx', data_only=True)
    sheet = chart['Sheet1']
    print("open file successfully")

    i = 2
    while i < 8640:
        if sheet.cell(row=i, column=12).value != 0:
            sheet.cell(row=i, column=12).value = 1

        if i % 500 == 0:
            print("计算到第%d行" % i)
            chart.save(os.path.dirname(__file__) + '/data/raw_data.xlsx')
        i += 1

    chart.save(os.path.dirname(__file__) + '/data/raw_data.xlsx')


def distributionAnalyze(column):
    s = pd.read_excel(os.path.dirname(__file__) + '/data/raw_data.xlsx', sheet_name='Sheet1', usecols=column)
    # scatter plot & histogram
    fig = plt.figure(figsize=(10, 6))
    ax1 = fig.add_subplot(2, 1, 1)  # 创建子图1
    ax1.scatter(s.index, s.values)
    plt.grid()


    ax2 = fig.add_subplot(2, 1, 2)  # 创建子图2
    s.hist(bins=30, alpha=0.5, ax=ax2)
    s.plot(kind='kde', secondary_y=True, ax=ax2)
    plt.grid()
    plt.savefig(os.path.dirname(__file__) + '\\image\\TotalCallDuration.png')


def FeatureSelection():
    chart = xlrd.open_workbook(os.path.dirname(__file__) + '/data/raw_data.xlsx')
    sheet = chart.sheet_by_name('Sheet1')

    # correlation maxtrix
    correlation = []
    for i in range(14):
        x = sheet.col_values(7+i)
        del x[0]
        j = 7+i
        correlation.append([])
        while j < 21:
            y = sheet.col_values(j)
            del y[0]

            # Calculate correlation between attributes via Spearman coefficient
            s, p = stats.spearmanr(x, y)
            correlation[i].append(s)

            j += 1

    print(correlation)

    chart = xl.load_workbook(os.path.dirname(__file__) + '/data/raw_data.xlsx', data_only=True)
    sheet_show = chart['Sheet3']

    for i in range(14):
        for j in range(14-i):
            sheet_show.cell(row=i+1, column=j+1).value = correlation[i][j]
            j += 1
        i += 1
    chart.save(os.path.dirname(__file__) + '/data/raw_data.xlsx')


def dataNormalize():
    chart = xl.load_workbook(os.path.dirname(__file__) + '/data/association_data.xlsx', data_only=True)
    sheet = chart['Sheet1']

    s = pd.read_excel(os.path.dirname(__file__) + '/data/association_data.xlsx', sheet_name='Sheet1', usecols='O')
    s.head()

    # min-max normalize
    min_max = preprocessing.MinMaxScaler()

    result = min_max.fit_transform(s)
    result = np.round(result, 5)  # 取5位小数

    j = 0
    while j in range(1):
        i = 0
        while i < 8638:
            sheet.cell(row=i+2, column=j+15).value = float(result[i][j])
            i += 1
        chart.save(os.path.dirname(__file__) + '/data/association_data.xlsx')
        j += 1


def dateTransform():
    chart = xl.load_workbook(os.path.dirname(__file__) + '/data/classify_data.xlsx', data_only=True)
    sheet = chart['Sheet1']

    i = 0
    while i < 8638:
        date = sheet.cell(row=i+2, column=8).value  # load data
        # Transfer strings e.g. “20210101” into dates
        date = time.mktime(time.strptime(date, '%Y%m%d'))
        now = time.time()  # time record
        total_s = now - date
        total_d = int(total_s/(60*60*24))  # count days
        sheet.cell(row=i+2, column=8).value = total_d  # write data
        i += 1
    chart.save(os.path.dirname(__file__) + '/data/classify_data.xlsx')


def CustomerLevel():
    raw_data = pd.read_excel(os.path.dirname(__file__) + '/data/raw_data.xlsx', sheet_name='Sheet1', usecols='A')
    raw_data.head()

    level_data = pd.read_excel(os.path.dirname(__file__) + '/data/attr_supplement.xlsx', sheet_name='Sheet1',
                               usecols='A, I')
    level_data.head()
    print('Open 2 files Successfully')

    # result = pd.DataFrame()
    result = []

    # drop duplicates for phone number
    raw_data.drop_duplicates(subset='PhoneNumber', keep='first', inplace=True)
    raw_data = raw_data.reset_index(drop=True)
    # raw_data = raw_data.tolist()

    for i in range(raw_data.shape[0]):
        # print(i)
        temp = level_data.where(level_data.PhoneNumber == raw_data['PhoneNumber'][i])
        # drop NA
        temp = temp.dropna()
        temp = temp.reset_index(drop=True)
        m, n = temp.shape

        # count repetition
        levels = [raw_data['PhoneNumber'][i], m]

        # record user level
        for j in range(m):
            s = temp['Level'][j]
            levels.extend([s])
        result.append(levels)

        if i % 1000 == 0:
            print('%s rows have been finished reading' % i)

    df = pd.DataFrame(result)
    df.to_excel(os.path.dirname(__file__) + '/data/level _data.xlsx')


if __name__ == '__main__':
    # Merge 2 tables
    # attrMerge()

    # Boolean attributes cleaning
    # boolMake()

    # Normal distribution dermination
    # distributionAnalyze('P')

    # Feature selection
    # FeatureSelection()

    # Data normalization
    # dataNormalize()

    # Transfer dates into count of days
    # dateTransform()

    # Custermers' level association
    CustomerLevel()
